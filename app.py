import streamlit as st
import pandas as pd
import xml.etree.ElementTree as ET
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO

# 1. Configuração da página
st.set_page_config(
    page_title="Conversor de XML em dados",
    layout="wide",
    page_icon="⚡"
)

# 2. Estilo Visual Otimizado (Dark Mode + Alto Contraste)
st.markdown("""
<style>
    .stApp {
        background-color: #0B0F19 !important;
        background-image: radial-gradient(circle at 10% 20%, rgba(0, 240, 255, 0.06) 0%, transparent 40%),
                          radial-gradient(circle at 90% 80%, rgba(138, 43, 226, 0.06) 0%, transparent 40%) !important;
        color: #FFFFFF !important;
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif !important;
    }

    section[data-testid="stSidebar"] {
        background-color: #0D131F !important;
        border-right: 1px solid rgba(255, 255, 255, 0.1) !important;
    }
    section[data-testid="stSidebar"] * {
        color: #FFFFFF !important;
    }

    p, span, label, div {
        color: #FFFFFF !important;
    }
    
    h1, h2, h3, h4, h5, h6 {
        color: #FFFFFF !important;
        font-weight: 700 !important;
    }

    div[data-testid="stRadio"] label, 
    div[data-testid="stRadio"] label p, 
    div[data-testid="stRadio"] span,
    div[data-testid="stCheckbox"] label, 
    div[data-testid="stCheckbox"] label p, 
    div[data-testid="stCheckbox"] span {
        color: #FFFFFF !important;
        font-size: 15px !important;
        font-weight: 500 !important;
    }

    .cyber-header {
        background: linear-gradient(135deg, rgba(15, 23, 42, 0.95), rgba(30, 41, 59, 0.85));
        border: 1px solid rgba(0, 240, 255, 0.35);
        border-radius: 16px;
        padding: 22px 28px;
        margin-bottom: 24px;
        box-shadow: 0 10px 30px rgba(0, 0, 0, 0.5);
    }
    .cyber-tag {
        color: #00F0FF !important;
        font-size: 12px;
        font-weight: 800;
        letter-spacing: 2px;
        text-transform: uppercase;
    }
    .cyber-title {
        color: #FFFFFF !important;
        font-size: 28px;
        font-weight: 800;
        margin: 4px 0 6px 0;
    }
    .cyber-desc {
        color: #CBD5E1 !important;
        font-size: 14px;
        margin: 0;
    }

    div[data-testid="stFileUploader"] {
        background: #111827 !important;
        border: 1px dashed rgba(0, 240, 255, 0.45) !important;
        border-radius: 14px !important;
        padding: 18px !important;
    }
    div[data-testid="stFileUploader"] * {
        color: #FFFFFF !important;
    }

    div.stButton > button:first-child {
        background: linear-gradient(135deg, #00F0FF 0%, #3B82F6 50%, #8B5CF6 100%) !important;
        color: #FFFFFF !important;
        font-weight: 700 !important;
        font-size: 16px !important;
        border: none !important;
        border-radius: 12px !important;
        padding: 14px 28px !important;
        box-shadow: 0 4px 20px rgba(0, 240, 255, 0.35) !important;
    }

    div.stDownloadButton > button:first-child {
        background: linear-gradient(135deg, #10B981 0%, #059669 100%) !important;
        color: #FFFFFF !important;
        font-weight: 700 !important;
        font-size: 16px !important;
        border: none !important;
        border-radius: 12px !important;
        padding: 14px 28px !important;
        box-shadow: 0 4px 20px rgba(16, 185, 129, 0.35) !important;
    }

    .kpi-container {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
        gap: 16px;
        margin: 20px 0;
    }
    .kpi-box {
        background: #111827;
        border: 1px solid rgba(0, 240, 255, 0.3);
        border-radius: 14px;
        padding: 16px 20px;
    }
    .kpi-label {
        font-size: 12px;
        color: #94A3B8 !important;
        text-transform: uppercase;
        letter-spacing: 1px;
        font-weight: 700;
    }
    .kpi-value {
        font-size: 24px;
        font-weight: 800;
        color: #00F0FF !important;
        margin-top: 4px;
    }
</style>
""", unsafe_allow_html=True)

# 3. Controle de Acesso por Senha
SENHA_DE_ACESSO = st.secrets.get("SENHA_ACESSO", "suasenha123")

def verificar_autenticacao():
    if "autenticado" not in st.session_state:
        st.session_state.autenticado = False

    if not st.session_state.autenticado:
        st.markdown("""
        <div style="max-width: 460px; margin: 50px auto 20px auto; background: #111827; border: 1px solid rgba(0, 240, 255, 0.35); border-radius: 18px; padding: 28px; text-align: center;">
            <div style="font-size: 36px; margin-bottom: 6px;">🔐</div>
            <div style="color: #00F0FF; font-size: 13px; letter-spacing: 2px; font-weight: 800;">ACESSO SEGURO</div>
            <h2 style="color: #FFF; margin: 6px 0 10px 0; font-size: 22px;">Conversor de XML em dados</h2>
            <p style="color: #CBD5E1; font-size: 13.5px; margin: 0;">Insira a credencial de segurança para acessar o sistema.</p>
        </div>
        """, unsafe_allow_html=True)
        
        col_esq, col_login, col_dir = st.columns(3)
        with col_login:
            senha = st.text_input("Credencial de Acesso", type="password", label_visibility="collapsed", placeholder="Digite sua senha...")
            if st.button("Autenticar e Entrar", use_container_width=True):
                if senha == SENHA_DE_ACESSO:
                    st.session_state.autenticado = True
                    st.rerun()
                else:
                    st.error("Credencial inválida. Acesso negado.")
        return False
    return True

if not verificar_autenticacao():
    st.stop()

# 4. Header Principal
st.markdown("""
<div class="cyber-header">
    <div>
        <span class="cyber-tag">⚡ INTELLIGENCE MATRIX V2.0</span>
        <div class="cyber-title">Conversor de XML em dados</div>
        <p class="cyber-desc">Extração automatizada de XMLs fiscais, consolidação analítica e exportação em planilha formatada.</p>
    </div>
</div>
""", unsafe_allow_html=True)

with st.sidebar:
    st.markdown("### 🛡️ Painel de Segurança")
    st.write("Sessão ativa e autenticada.")
    if st.button("Encerrar Sessão", use_container_width=True):
        st.session_state.autenticado = False
        st.rerun()

# 5. Upload de Arquivos
st.markdown("### 📂 1. Importação de Arquivos XML")
st.markdown("<p style='color: #CBD5E1; font-size: 14px;'>Selecione ou arraste os arquivos XML das notas fiscais (modelos 55, 65 e 59):</p>", unsafe_allow_html=True)
arquivos_xml = st.file_uploader("Upload de XMLs", type=["xml"], accept_multiple_files=True, label_visibility="collapsed")

st.markdown("<br>", unsafe_allow_html=True)

# 6. Configurações de Consolidação e Colunas
st.markdown("### ⚙️ 2. Parâmetros de Consolidação e Colunas")
col_card1, col_card2 = st.columns(2)

with col_card1:
    st.markdown("""
    <div style="background: #111827; border: 1px solid rgba(0, 240, 255, 0.25); border-radius: 12px; padding: 14px 18px; margin-bottom: 12px;">
        <span style="color: #00F0FF; font-weight: 700; font-size: 13.5px; letter-spacing: 1px;">📊 REGRA DE DUPLICADOS</span>
        <p style="color: #CBD5E1; font-size: 12.5px; margin: 4px 0 0 0;">Como consolidar itens repetidos ao longo dos meses:</p>
    </div>
    """, unsafe_allow_html=True)
    modo_consolidacao = st.radio(
        "Modo de consolidação:",
        options=[
            "Consolidar produtos (Somar quantidades e calcular custo médio)",
            "Manter apenas o preço da última compra (Custo mais recente)",
            "Listar todas as compras (Histórico completo sem agrupar)"
        ],
        index=0,
        label_visibility="collapsed"
    )

with col_card2:
    st.markdown("""
    <div style="background: #111827; border: 1px solid rgba(139, 92, 246, 0.3); border-radius: 12px; padding: 14px 18px; margin-bottom: 12px;">
        <span style="color: #C084FC; font-weight: 700; font-size: 13.5px; letter-spacing: 1px;">📌 SELEÇÃO DE CAMPOS EXPORTADOS</span>
        <p style="color: #CBD5E1; font-size: 12.5px; margin: 4px 0 0 0;">Marque as colunas que devem constar no relatório:</p>
    </div>
    """, unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        f_estab = st.checkbox("Estabelecimento", value=True)
        f_end = st.checkbox("Endereço", value=True)
        f_cod = st.checkbox("Código do Produto", value=True)
        f_ean = st.checkbox("EAN", value=True)
        f_desc = st.checkbox("Descrição do Produto", value=True)
    with c2:
        f_qtd = st.checkbox("Quantidade", value=True)
        f_unit = st.checkbox("Custo Unitário", value=True)
        f_total = st.checkbox("Custo Total", value=True)
        f_data = st.checkbox("Data", value=True)
        f_arquivo = st.checkbox("Arquivo", value=True)

# Lista de colunas marcadas
colunas_selecionadas = []
if f_estab: colunas_selecionadas.append("Estabelecimento")
if f_end: colunas_selecionadas.append("Endereço")
if f_cod: colunas_selecionadas.append("Código do Produto")
if f_ean: colunas_selecionadas.append("EAN")
if f_desc: colunas_selecionadas.append("Descrição do Produto")
if f_qtd: colunas_selecionadas.append("Quantidade")
if f_unit: colunas_selecionadas.append("Custo Unitário")
if f_total: colunas_selecionadas.append("Custo Total")
if f_data: colunas_selecionadas.append("Data")
if f_arquivo: colunas_selecionadas.append("Arquivo")

# 7. Funções de Processamento de XML
def parse_xml_content(content_bytes, file_name=""):
    try:
        text = content_bytes.decode('utf-8')
    except UnicodeDecodeError:
        try:
            text = content_bytes.decode('iso-8859-1')
        except Exception:
            text = content_bytes.decode('utf-8', errors='ignore')
            
    clean_text = re.sub(r'<(/?)(\w+):', r'<\1', text)
    clean_text = re.sub(r'\sxmlns(:\w+)?="[^"]+"', '', clean_text)
    root = ET.fromstring(clean_text)

    # 1. Dados da NF (Data e Número)
    numero_nf = ''
    data_emissao = ''
    data_iso = ''
    
    nNF_el = root.find('.//nNF') or root.find('.//nCFe') or root.find('.//nDoc')
    if nNF_el is not None and nNF_el.text:
        numero_nf = nNF_el.text.strip()
        
    for tag in ['dhEmi', 'dEmi', 'dhSaiEnt', 'dSaiEnt', 'dhRecbto', 'dCompet']:
        el = root.find(f'.//{tag}')
        if el is not None and el.text and el.text.strip():
            raw = el.text.strip()
            if '-' in raw and len(raw) >= 10:
                iso_part = raw[:10]
                p = iso_part.split('-')
                if len(p) == 3 and len(p[0]) == 4:
                    data_emissao = f"{p}/{p}/{p[0]}"
                    data_iso = iso_part
                    break
            elif len(raw) == 8 and raw.isdigit():
                yyyy, mm, dd = raw[:4], raw[4:6], raw[6:8]
                data_emissao = f"{dd}/{mm}/{yyyy}"
                data_iso = f"{yyyy}-{mm}-{dd}"
                break
                
    if not data_emissao:
        match = re.search(r'<(?:dhEmi|dEmi|dhSaiEnt|dhRecbto)[^>]*>(\d{4}-\d{2}-\d{2})', text)
        if match:
            iso_part = match.group(1)
            p = iso_part.split('-')
            data_emissao = f"{p}/{p}/{p[0]}"
            data_iso = iso_part
        
    # 2. Fornecedor
    nome_fornecedor = ''
    endereco_fornecedor = ''
    emit = root.find('.//emit')
    if emit is not None:
        xNome = emit.find('xNome')
        nome_fornecedor = xNome.text.strip() if xNome is not None and xNome.text else ''
        
        ender = emit.find('enderEmit')
        if ender is not None:
            def get_field(tag):
                el = ender.find(tag)
                return el.text.strip() if el is not None and el.text else ''
            
            lgr = get_field('xLgr')
            nro = get_field('nro')
            bairro = get_field('xBairro')
            mun = get_field('xMun')
            uf = get_field('UF')
            endereco_fornecedor = f"{lgr}, {nro} - {bairro}, {mun} - {uf}".strip(" ,-")
            
    # 3. Itens
    dets = root.findall('.//det')
    itens = []
    for det in dets:
        prod = det.find('prod')
        if prod is None:
            continue
            
        def get_val(tag, default=''):
            el = prod.find(tag)
            return el.text.strip() if el is not None and el.text else default
            
        cProd = get_val('cProd')
        cEAN = get_val('cEAN')
        xProd = get_val('xProd')
        qCom_str = get_val('qCom', '0')
        vUnCom_str = get_val('vUnCom', '0')
        vProd_str = get_val('vProd', '0')
        
        try:
            qCom = float(qCom_str)
        except ValueError:
            qCom = 0.0
            
        try:
            vUnCom = float(vUnCom_str)
        except ValueError:
            vUnCom = 0.0
            
        try:
            vProd = float(vProd_str)
        except ValueError:
            vProd = 0.0
            
        itens.append({
            'Estabelecimento': nome_fornecedor,
            'Endereço': endereco_fornecedor,
            'Código do Produto': cProd,
            'EAN': cEAN,
            'Descrição do Produto': xProd,
            'Quantidade': qCom,
            'Custo Unitário': vUnCom,
            'Custo Total': vProd,
            'Data': data_emissao,
            'Data_ISO': data_iso,
            'Número da NF': numero_nf,
            'Arquivo': file_name
        })
        
    return itens

def processar_todos_xmls(arquivos):
    todos_itens = []
    arquivos_com_erro = []
    
    for arquivo in arquivos:
        try:
            conteudo = arquivo.read()
            itens = parse_xml_content(conteudo, file_name=arquivo.name)
            if itens:
                todos_itens.extend(itens)
            else:
                arquivos_com_erro.append(f"{arquivo.name} (nenhum produto encontrado)")
        except Exception as e:
            arquivos_com_erro.append(f"{arquivo.name} (erro: {e})")
            
    return pd.DataFrame(todos_itens), arquivos_com_erro

def aplicar_tratamento_e_colunas(df, modo, colunas_escolhidas):
    if df.empty:
        return df

    if 'Data_ISO' in df.columns:
        df = df.sort_values(by=['Data_ISO', 'Número da NF'])

    # Agrupa exclusivamente pelo PRODUTO (eliminando duplicados entre filiais/lojas)
    group_cols = ['Código do Produto', 'EAN', 'Descrição do Produto']

    if modo.startswith("Consolidar produtos"):
        df_agrupado = df.groupby(group_cols, as_index=False).agg(
            Estabelecimento=('Estabelecimento', 'last'),
            Endereço=('Endereço', 'last'),
            Quantidade=('Quantidade', 'sum'),
            Custo_Total=('Custo Total', 'sum'),
            Ultimo_Custo=('Custo Unitário', 'last'),
            Data=('Data', 'last'),
            Arquivo=('Arquivo', lambda x: ', '.join(x.unique()))
        )
        df_agrupado['Custo Unitário'] = (df_agrupado['Custo_Total'] / df_agrupado['Quantidade']).round(2)
        df_agrupado['Custo Total'] = df_agrupado['Custo_Total'].round(2)
        df_final = df_agrupado

    elif modo.startswith("Manter apenas o preço da última compra"):
        df_final = df.drop_duplicates(subset=group_cols, keep='last').copy()

    else:
        df_final = df.copy()

    # Ordem das Colunas solicitada:
    # Estabelecimento, Endereço, Código do Produto, EAN, Descrição do Produto, Quantidade, Custo Unitário, Custo Total (Antepenúltima), Data (Penúltima), Arquivo (Última)
    ordem_padrao = [
        'Estabelecimento', 'Endereço', 'Código do Produto', 'EAN', 
        'Descrição do Produto', 'Quantidade', 'Custo Unitário', 
        'Custo Total', 'Data', 'Arquivo'
    ]

    colunas_finais = [c for c in ordem_padrao if c in colunas_escolhidas and c in df_final.columns]
    return df_final[colunas_finais]

def gerar_excel(df, titulo_aba="Notas Fiscais"):
    wb = Workbook()
    ws = wb.active
    ws.title = titulo_aba[:31]
    headers = list(df.columns)
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = align_center
        c.border = border

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid") if r_idx % 2 == 0 else PatternFill(fill_type=None)
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.fill = fill
            cell.border = border
            header = headers[c_idx - 1]
            
            if 'Quantidade' in header:
                cell.number_format = '#,##0.00'
                cell.alignment = align_center
            elif 'Custo' in header or 'Preço' in header or 'Valor' in header:
                cell.number_format = '"R$ "#,##0.00'
                cell.alignment = align_center
            elif header in ['Código do Produto', 'EAN', 'Arquivo', 'Data']:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column].width = min(max_length + 3, 60)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

# 8. Execução
st.markdown("<br>", unsafe_allow_html=True)

if arquivos_xml:
    if not colunas_selecionadas:
        st.warning("⚠️ Marque ao menos uma coluna acima para gerar o relatório.")
    else:
        if st.button("⚡ Executar Processamento e Gerar Planilha", type="primary", use_container_width=True):
            with st.spinner("Processando dados e estruturando planilha..."):
                df_bruto, erros = processar_todos_xmls(arquivos_xml)
                
                if not df_bruto.empty:
                    df_final = aplicar_tratamento_e_colunas(df_bruto, modo_consolidacao, colunas_selecionadas)
                    
                    total_arquivos = len(arquivos_xml)
                    total_compras = len(df_bruto)
                    total_linhas = len(df_final)
                    valor_total = df_bruto['Custo Total'].sum() if 'Custo Total' in df_bruto.columns else 0.0
                    
                    st.markdown(f"""
                    <div class="kpi-container">
                        <div class="kpi-box">
                            <div class="kpi-label">📁 Arquivos Lidos</div>
                            <div class="kpi-value">{total_arquivos}</div>
                        </div>
                        <div class="kpi-box">
                            <div class="kpi-label">📦 Total de Compras</div>
                            <div class="kpi-value">{total_compras}</div>
                        </div>
                        <div class="kpi-box">
                            <div class="kpi-label">💎 Produtos no Relatório</div>
                            <div class="kpi-value">{total_linhas}</div>
                        </div>
                        <div class="kpi-box">
                            <div class="kpi-label">💰 Valor Consolidado</div>
                            <div class="kpi-value">R$ {valor_total:,.2f}</div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    if erros:
                        with st.expander("⚠️ Alertas de arquivos com inconsistência:"):
                            for err in erros:
                                st.warning(err)
                    
                    excel_data = gerar_excel(df_final, titulo_aba="Relatório NFe")
                    
                    st.download_button(
                        label="📥 Baixar Planilha Excel Formatada (.xlsx)",
                        data=excel_data,
                        file_name="Relatorio_Notas_Fiscais.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    
                    st.markdown("#### 🔍 Prévia dos Dados Estruturados")
                    st.dataframe(df_final, use_container_width=True)
                else:
                    st.error("Nenhum produto válido foi identificado nos arquivos enviados.")
                    if erros:
                        for err in erros:
                            st.warning(err)
